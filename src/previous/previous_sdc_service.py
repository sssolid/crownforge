"""
SDC Template Service
Handles SDC template population and processing
"""

import logging
import json
import re
import unicodedata
from pathlib import Path
from typing import List, Dict, Any, Optional, Set
from dataclasses import dataclass
from openpyxl import load_workbook

from .database import FilemakerService
from .utils import performance_monitor, TextProcessor

logger = logging.getLogger(__name__)


@dataclass
class SDCTemplateConfig:
    """SDC template configuration"""
    country_mapping_file: str
    packaging_mapping_file: str
    interchange_codes: List[str]
    field_limits: Dict[str, int]


class SDCService:
    """Service for SDC template operations"""

    def __init__(self, filemaker_service: FilemakerService, config: Dict[str, Any]):
        self.filemaker = filemaker_service
        self.config = SDCTemplateConfig(**config)
        self.country_map = self._load_country_mapping()
        self.packaging_map = self._load_packaging_mapping()
        self.popularity_verbiage = self._get_popularity_verbiage()

    def _load_country_mapping(self) -> Dict[str, str]:
        """Load country code mapping"""
        try:
            with open(self.config.country_mapping_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.warning(f"Country mapping file not found: {self.config.country_mapping_file}")
            return self._get_default_country_mapping()

    def _load_packaging_mapping(self) -> Dict[str, str]:
        """Load packaging code mapping"""
        try:
            with open(self.config.packaging_mapping_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.warning(f"Packaging mapping file not found: {self.config.packaging_mapping_file}")
            return self._get_default_packaging_mapping()

    def _get_default_country_mapping(self) -> Dict[str, str]:
        """Get default country code mapping"""
        return {
            "ARGENTINA": "AR",
            "ASSEMBLED IN USA": "US",
            "ASSEMBLED IN USA WITH COMPONENTS FROM BRAZIL & MEXICO": "BR",
            "ASSEMBLED IN USA WITH COMPONENTS FROM CHINA AND USA": "CN",
            "Assembled in USA with components from CHINA, USA": "CN",
            "Assembled in USA with components from PAKISTAN, USA": "PK",
            "ASSEMBLED IN USA WITH COMPONENTS FROM USA & PORTUGAL": "PT",
            "ASSEMBLED IN USA WITH COMPONENTS FROM USA and INDIA": "IN",
            "ASSEMBLED IN USA WITH COMPONENTS FROM USA AND PAKISTAN": "PK",
            "ASSEMBLED IN USA WITH COMPONENTS FROM VARIOUS COUNTRIES": "US",
            "BRAZIL": "BR",
            "BRAZIL & MEXICO": "BR",
            "CANADA": "CA",
            "cH": "CH",
            "CHINA": "CN",
            "CHINA & INDIA": "CN",
            "CHINA & TAIWAN": "TW",
            "CHINA & USA": "CN",
            "CHINA AND USA": "CN",
            "CHINAUSA": "CN",
            "CHINA, BRAZIL": "BR",
            "CHINAAIWAN": "CN",
            "COLOMBIA": "CO",
            "COMPONENTS FROM CHINA & TAIWAN": "CN",
            "COMPONENTS FROM CHINA & USA": "CN",
            "Components from China, Taiwan": "CN",
            "COMPONENTS FROM CHINA, USA & TAIWAN": "TW",
            "COMPONENTS FROM CHINA, USA, TAIWAN": "TW",
            "COMPONENTS FROM INDIA AND USA": "IN",
            "COMPONENTS FROM PAKISTAN AND USA": "PK",
            "COMPONENTS FROM TAIWAN & USA": "TW",
            "COMPONENTS FROM TAIWAN AND CHINA": "TW",
            "COMPONENTS MADE IN CHINA": "CN",
            "COMPONENTS MADE IN CHINA AND TAIWAN": "CN",
            "COMPONENTS MADE IN CHINA AND USA": "CN",
            "COMPONENTS MADE IN TAIWAN, USA CHINA": "CN",
            "COMPONENTS MADE IN USA AND CHINA": "CN",
            "COMPONENTS MADE IN USA AND TAIWAN": "TW",
            "COMPONENTS MADE IN USA, TAIWAN AND CHINA": "TW",
            "CONTAINS COMPONENTS FROM CHINA AND TAIWAN": "CN",
            "Contains components from Taiwan, USA": "TW",
            "CZECH REPUBLIC": "CZ",
            "FRANCE": "FR",
            "GERMANY": "DE",
            "INCLUDES COMPONENTS FROM TAIWAN AND USA": "TW",
            "INCLUDES COMPONENTS MADE IN CHINA": "CN",
            "INCLUDES COMPONENTS MADE IN CHINA, INDIA": "IN",
            "INCLUDES COMPONENTS MADE IN INDIA": "IN",
            "INCLUDES COMPONENTS MADE IN INDIA, CHINA": "IN",
            "INDIA": "IN",
            "INDIA & USA": "IN",
            "INDONESIA": "ID",
            "ISRAEL": "IL",
            "ISRAEL & USA": "IL",
            "ISRAEL, CHINA & USA": "CN",
            "ITALY": "IT",
            "JAPAN": "JP",
            "JAPAN & TAIWAN": "TW",
            "KOREA": "KR",
            "MEXICO": "MX",
            "PAKISTAN": "PK",
            "PAKISTAN & CHINA": "PK",
            "PERU": "PE",
            "POLAND": "PL",
            "PORTUGAL": "PT",
            "S. AFRICA": "ZA",
            "SLOVAKIA": "SK",
            "SOUTH KOREA": "KR",
            "SPAIN": "ES",
            "TAIWAN": "TW",
            "TAIWAN & INDIA": "TW",
            "TAIWAN & USA": "TW",
            "THAILAND": "TH",
            "TURKEY": "TR",
            "USA": "US",
            "USA & CHINA": "CN",
            "USA & INDIA": "IN",
            "USA AND CHINA": "CN",
            "USA AND TAIWAN": "TW",
            "USA, TAIWAN, & CHINA": "CN",
            "USA,UNITED KINGDOM": "GB",
            "VARIOUS COUNTRIES": "US",
            "VIETNAM": "VN",
        }

    def _get_default_packaging_mapping(self) -> Dict[str, str]:
        """Get default packaging code mapping"""
        return {
            'Box': 'BX',
            'Poly Bag': 'BG',
            'Loose': 'BX',
            'Bubble Wrapped': 'BX',
            'Blister Pack': 'PK',
            'Shrink Wrap on Cardboard': 'PK',
            'Sleeve': 'SL',
            'Bottle': 'BO',
            'Ask': 'BX'
        }

    def _get_popularity_verbiage(self) -> Dict[str, str]:
        """Get popularity code descriptions"""
        return {
            'A': 'Top 60% of Product Group Sales Value (Units x $, Hits, etc.)',
            'B': 'Next 20% of Product Group Sales Value',
            'C': 'Next 15% of Product Group Sales Value',
            'D': 'Last 5% of Product Group Sales Value',
            'N': 'New Item 3 Months, 6 Months, etc.',
        }

    def sanitize_string(self, value: Any) -> str:
        """Sanitize string for Excel output"""
        if value and isinstance(value, str):
            # Normalize Unicode to decompose formatting marks
            value = unicodedata.normalize('NFKD', value)
            # Remove all control characters, invisible characters, and strip whitespace
            value = re.sub(r'[\x00-\x1F\x7F-\x9F\u2000-\u200F\u2028-\u202F\u205F-\u206F\xa0]', '', value)
            # Remove Trademark symbol (™) and Registered Trademark symbol (®)
            value = value.replace('™', '').replace('®', '')
            # Return the stripped and sanitized value
            return value.strip()
        return value if value is not None else ''

    def load_missing_part_numbers(self, missing_parts_file: str) -> Set[str]:
        """Load missing part numbers from text file"""
        missing_parts = set()

        try:
            with open(missing_parts_file, 'r') as f:
                for line in f:
                    part_number = line.strip()
                    if part_number:
                        missing_parts.add(part_number)

            logger.info(f"Loaded {len(missing_parts)} missing part numbers")

        except FileNotFoundError:
            logger.warning(f"Missing parts file not found: {missing_parts_file}")

        return missing_parts

    def populate_sdc_template(self, template_file: str, output_file: str,
                              missing_parts_file: Optional[str] = None,
                              popularity_mapping: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
        """
        Populate SDC template with product data

        Args:
            template_file: Path to blank SDC template
            output_file: Path for populated template output
            missing_parts_file: Optional file with specific part numbers to include
            popularity_mapping: Optional mapping of part numbers to popularity codes

        Returns:
            Dictionary with population results
        """
        logger.info(f"Populating SDC template: {template_file} -> {output_file}")

        with performance_monitor("SDC Template Population") as monitor:
            # Load missing part numbers if specified
            missing_parts = set()
            if missing_parts_file:
                missing_parts = self.load_missing_part_numbers(missing_parts_file)

            # Get product data from Filemaker
            product_data = self.filemaker.get_sdc_template_data(
                list(missing_parts) if missing_parts else None
            )
            monitor.increment_processed(len(product_data))

            # Get interchange data
            interchange_data = self.filemaker.get_interchange_data()
            monitor.increment_processed(len(interchange_data))

            # Load Excel template
            wb = load_workbook(filename=template_file)

            # Populate each sheet
            results = {}
            results['pies_item'] = self._populate_pies_item(wb, product_data)
            results['pies_descriptions'] = self._populate_pies_descriptions(wb, product_data)
            results['pies_prices'] = self._populate_pies_prices(wb, product_data)
            results['pies_extended_info'] = self._populate_pies_extended_info(wb, product_data, popularity_mapping)
            results['pies_packages'] = self._populate_pies_packages(wb, product_data)
            results['pies_user_attr'] = self._populate_pies_user_attr(wb, product_data)
            results['interchange'] = self._populate_interchange(wb, product_data, interchange_data)

            # Save populated template
            Path(output_file).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)

            summary = {
                'input_file': template_file,
                'output_file': output_file,
                'total_products': len(product_data),
                'sheets_populated': len(results),
                'missing_parts_filtered': len(missing_parts) > 0,
                'popularity_codes_applied': popularity_mapping is not None
            }

            logger.info(f"SDC template populated successfully: {summary}")
            return summary

    def _populate_pies_item(self, wb, product_data: List[Dict]) -> Dict[str, Any]:
        """Populate PiesItem sheet"""
        ws = wb['PiesItem']
        row_num = 2
        processed_count = 0

        for record in product_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue

            row_num += 1
            processed_count += 1

            try:
                ws[f"A{row_num}"] = 'Default'
                ws[f"B{row_num}"] = record.get('SDC_PartType', '')
                ws[f"C{row_num}"] = record.get('SDC_PartTerminologyID', '')
                ws[f"D{row_num}"] = record.get('AS400_NumberStripped', '')
                ws[f"E{row_num}"] = 'Y' if record.get('PartHazardousMaterial') == 'Yes' else 'N'

                upc = record.get('AS400_UPC')
                ws[f"G{row_num}"] = '00' + str(upc) if upc else ''
                ws[f"H{row_num}"] = 'UP'
                ws[f"I{row_num}"] = 'GGMQ' if record.get('PartBrand') == 'RT Off-Road' else 'BKMN'
                ws[f"J{row_num}"] = 'RT Off-Road' if record.get(
                    'PartBrand') == 'RT Off-Road' else 'Crown Automotive Jeep Replacement'

                if record.get('PartApplication') == 'Universal':
                    ws[f"M{row_num}"] = 'N'
                else:
                    ws[f"M{row_num}"] = 'Y'

                ws[f"N{row_num}"] = record.get('PartQuantityInPackage', '') or ''
                ws[f"O{row_num}"] = 'EA'

                packaging = record.get('PartPackaging', '')
                ws[f"P{row_num}"] = self.packaging_map.get(packaging, '')

                # Handle quantity required logic
                qty_required = str(record.get('PartQuantityRequired', '')).strip()
                if qty_required == 'Varies with Application':
                    ws[f"Q{row_num}"] = 'MAX'
                    ws[f"R{row_num}"] = '1'
                else:
                    qty_in_package = record.get('PartQuantityInPackage')
                    qty_required_num = record.get('PartQuantityRequired')

                    if qty_in_package and qty_required_num:
                        try:
                            qty_in = float(qty_in_package)
                            qty_req = float(qty_required_num)

                            if qty_in < qty_req:
                                ws[f"Q{row_num}"] = 'MIN'
                            else:
                                ws[f"Q{row_num}"] = 'NOR'

                            ws[f"R{row_num}"] = str(qty_required_num)
                        except ValueError:
                            ws[f"Q{row_num}"] = 'MIN'
                            ws[f"R{row_num}"] = qty_required or ''
                    else:
                        ws[f"Q{row_num}"] = 'MIN'
                        ws[f"R{row_num}"] = qty_required or ''

                ws[f"S{row_num}"] = 'EA'
                ws[f"V{row_num}"] = '1'
                ws[f"W{row_num}"] = 'EA'

            except Exception as e:
                logger.error(f"Error populating PiesItem row {row_num}: {e}")
                continue

        return {'processed_count': processed_count}

    def _populate_pies_descriptions(self, wb, product_data: List[Dict]) -> Dict[str, Any]:
        """Populate PiesDescriptions sheet"""
        ws = wb['PiesDescriptions']
        row_num = 2
        processed_count = 0

        for record in product_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue

            row_num += 1
            processed_count += 1

            try:
                ws[f"A{row_num}"] = 'Default'
                ws[f"B{row_num}"] = record.get('SDC_PartType', '')
                ws[f"C{row_num}"] = record.get('AS400_NumberStripped', '')
                ws[f"D{row_num}"] = 'EN'

                # Handle description fields with length limits
                limits = self.config.field_limits

                # Abbreviated Description (12 chars)
                desc_abbrev = record.get('SDC_DescriptionAbbreviated') or record.get('PartDescription', '')
                ws[f"E{row_num}"] = str(desc_abbrev)[:limits.get('abbreviated_description', 12)] if desc_abbrev else ''

                # Short Description (20 chars)
                desc_short = record.get('SDC_DescriptionShort') or record.get('PartDescription', '')
                ws[f"F{row_num}"] = str(desc_short)[:limits.get('short_description', 20)] if desc_short else ''

                # Invoice Description (40 chars)
                desc_invoice = record.get('SDC_DescriptionInvoice') or record.get('PartDescription', '')
                ws[f"G{row_num}"] = str(desc_invoice)[:limits.get('invoice_description', 40)] if desc_invoice else ''

                ws[f"I{row_num}"] = record.get('SDC_KeySearchWords', '')

                # Long descriptions
                long_desc = record.get('SDC_PartDescriptionLongJeepOnly') or record.get('PartDescriptionLongJeepOnly',
                                                                                        '')
                ws[f"J{row_num}"] = long_desc
                ws[f"K{row_num}"] = long_desc

                ws[f"M{row_num}"] = record.get('SDC_SlangDescription', '')
                ws[f"N{row_num}"] = record.get('PartApplication', '')

                extended_desc = record.get('SDC_PartDescriptionExtended') or record.get('PartDescriptionExtended', '')
                ws[f"O{row_num}"] = extended_desc

                # RT Off-Road specific fields
                ws[f"P{row_num}"] = record.get('RTOffRoadAdCopy', '')

                # Combine bullet points
                bullet_keys = [f'RTOffRoadBullet{i}' for i in range(1, 12)]
                valid_bullets = []

                for key in bullet_keys:
                    bullet = record.get(key, '')
                    if bullet and len(bullet) <= limits.get('bullet_points_max', 240):
                        valid_bullets.append(bullet)

                ws[f"S{row_num}"] = ";".join(valid_bullets)
                ws[f"U{row_num}"] = record.get('SDC_AAIAPartTypeDescription', '')

                # Brand-specific description
                brand = record.get('PartBrand', '')
                desc = record.get('PartDescription', '')
                if brand == 'RT Off-Road':
                    ws[f"W{row_num}"] = f'RT Off-Road {desc}'
                else:
                    ws[f"W{row_num}"] = f'Crown Automotive {desc}'

            except Exception as e:
                logger.error(f"Error populating PiesDescriptions row {row_num}: {e}")
                continue

        return {'processed_count': processed_count}

    def _populate_pies_prices(self, wb, product_data: List[Dict]) -> Dict[str, Any]:
        """Populate PiesPrices sheet"""
        ws = wb['PiesPrices']
        row_num = 2
        processed_count = 0

        for record in product_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue

            # Jobber price row
            row_num += 1
            try:
                ws[f"A{row_num}"] = 'Default'
                ws[f"B{row_num}"] = record.get('SDC_PartType', '')
                ws[f"C{row_num}"] = record.get('AS400_NumberStripped', '')
                ws[f"D{row_num}"] = 'J15'
                ws[f"E{row_num}"] = 'USD'
                ws[f"F{row_num}"] = '05/13/2025'
                ws[f"H{row_num}"] = 'JBR'
                ws[f"I{row_num}"] = record.get('AS400_JobberPrice', '') or ''
                ws[f"J{row_num}"] = 'PE'
            except Exception as e:
                logger.error(f"Error populating PiesPrices jobber row {row_num}: {e}")

            # Retail price row (jobber * 1.1)
            row_num += 1
            try:
                ws[f"A{row_num}"] = 'Default'
                ws[f"B{row_num}"] = record.get('SDC_PartType', '')
                ws[f"C{row_num}"] = record.get('AS400_NumberStripped', '')
                ws[f"D{row_num}"] = 'J15'
                ws[f"E{row_num}"] = 'USD'
                ws[f"F{row_num}"] = '09/09/2024'
                ws[f"H{row_num}"] = 'RET'

                jobber_price = record.get('AS400_JobberPrice')
                if jobber_price:
                    ws[f"I{row_num}"] = float(jobber_price) * 1.1
                else:
                    ws[f"I{row_num}"] = ''

                ws[f"J{row_num}"] = 'PE'
                processed_count += 1
            except Exception as e:
                logger.error(f"Error populating PiesPrices retail row {row_num}: {e}")

        return {'processed_count': processed_count}

    def _populate_pies_extended_info(self, wb, product_data: List[Dict],
                                     popularity_mapping: Optional[Dict[str, str]]) -> Dict[str, Any]:
        """Populate PiesExtendedInfo sheet"""
        ws = wb['PiesExtendedInfo']
        row_num = 2
        processed_count = 0

        # Normalize country mapping keys for case-insensitive lookup
        country_map_normalized = {key.lower(): value for key, value in self.country_map.items()}

        for record in product_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue

            row_num += 1
            processed_count += 1

            try:
                ws[f"A{row_num}"] = 'Default'
                ws[f"B{row_num}"] = record.get('SDC_PartType', '')
                ws[f"C{row_num}"] = record.get('AS400_NumberStripped', '')
                ws[f"D{row_num}"] = 'EN'
                ws[f"I{row_num}"] = '1'
                ws[f"K{row_num}"] = 'N'  # Hazardous material (disabled for now)
                ws[f"L{row_num}"] = 'Y'
                ws[f"M{row_num}"] = 'N'
                ws[f"P{row_num}"] = 'N'
                ws[f"Q{row_num}"] = 'N'
                ws[f"W{row_num}"] = '12'
                ws[f"X{row_num}"] = 'MO'

                # Country mapping
                part_country = record.get('PartCountry', '').strip().lower()
                if part_country and part_country in country_map_normalized:
                    ws[f"AC{row_num}"] = country_map_normalized[part_country]
                elif part_country:
                    logger.debug(
                        f"No country mapping for: {record.get('AS400_NumberStripped')} - {record.get('PartCountry')}")

                # Tariff code
                tariff = record.get('PartTariff', '')
                if tariff:
                    ws[f"AJ{row_num}"] = str(tariff).replace('.', '')

                ws[f"AN{row_num}"] = '2'
                ws[f"AO{row_num}"] = 'Available to Order'

                # Popularity code
                if popularity_mapping:
                    part_number = record.get('AS400_NumberStripped', '')
                    popularity_code = popularity_mapping.get(part_number, 'D')
                    ws[f"AP{row_num}"] = popularity_code

                ws[f"AX{row_num}"] = 'NEW'

            except Exception as e:
                logger.error(f"Error populating PiesExtendedInfo row {row_num}: {e}")
                continue

        return {'processed_count': processed_count}

    def _populate_pies_packages(self, wb, product_data: List[Dict]) -> Dict[str, Any]:
        """Populate PiesPackages sheet"""
        ws = wb['PiesPackages']
        row_num = 2
        processed_count = 0

        for record in product_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue

            row_num += 1
            processed_count += 1

            try:
                ws[f"A{row_num}"] = 'Default'
                ws[f"B{row_num}"] = record.get('SDC_PartType', '')
                ws[f"C{row_num}"] = record.get('AS400_NumberStripped', '')

                upc = record.get('AS400_UPC')
                upc_formatted = '00' + str(upc) if upc else ''
                ws[f"D{row_num}"] = upc_formatted
                ws[f"F{row_num}"] = upc_formatted

                ws[f"G{row_num}"] = 'EA'
                ws[f"H{row_num}"] = '1'
                ws[f"L{row_num}"] = record.get('AS400_Height', '')
                ws[f"M{row_num}"] = record.get('AS400_Width', '')
                ws[f"N{row_num}"] = record.get('AS400_Length', '')
                ws[f"O{row_num}"] = 'IN'
                ws[f"P{row_num}"] = record.get('AS400_Weight', '')
                ws[f"Q{row_num}"] = 'PG'
                ws[f"S{row_num}"] = record.get('PartDimensionalWeight', '')

            except Exception as e:
                logger.error(f"Error populating PiesPackages row {row_num}: {e}")
                continue

        return {'processed_count': processed_count}

    def _populate_pies_user_attr(self, wb, product_data: List[Dict]) -> Dict[str, Any]:
        """Populate PiesUserAttr sheet"""
        ws = wb['PiesUserAttr']
        row_num = 2
        processed_count = 0

        for record in product_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue

            record_sequence = 1
            part_number = record.get('AS400_NumberStripped', '')
            part_type = record.get('SDC_PartType', '')

            # Helper function to add attribute row
            def add_attribute_row(attribute_name: str, value: str):
                nonlocal row_num, record_sequence

                if value and value != 'N/A':
                    row_num += 1
                    ws[f"A{row_num}"] = 'Default'
                    ws[f"B{row_num}"] = part_type
                    ws[f"C{row_num}"] = part_number
                    ws[f"D{row_num}"] = attribute_name
                    ws[f"E{row_num}"] = 'N'

                    # Clean multiline values
                    if '\n' in str(value):
                        cleaned_value = ', '.join(line.strip() for line in str(value).splitlines() if line.strip())
                        ws[f"F{row_num}"] = cleaned_value
                    else:
                        ws[f"F{row_num}"] = str(value)

                    ws[f"G{row_num}"] = record_sequence
                    record_sequence += 1

            # Add various attributes
            # Brand-specific description
            brand = record.get('PartBrand', '')
            desc = record.get('PartDescription', '')
            if brand == 'RT Off-Road':
                add_attribute_row('Title', f'RT Off-Road {desc}')
            else:
                add_attribute_row('Title', f'Crown Automotive {desc}')
            add_attribute_row('Construction', record.get('PartConstruction', ''))
            add_attribute_row('Color', record.get('PartColor', ''))
            add_attribute_row('Texture', record.get('PartTexture', ''))
            add_attribute_row('Hardware', record.get('PartHardware', ''))
            add_attribute_row('Front/Rear', record.get('VehicleFrontRear', ''))
            add_attribute_row('Left/Right', record.get('VehicleLeftRight', ''))
            add_attribute_row('Upper/Lower', record.get('VehicleUpperLower', ''))
            add_attribute_row('Inner/Outer', record.get('VehicleInnerOuter', ''))

            # Additional shipping
            if record.get('PartAdditionalShipping') == 'Yes':
                add_attribute_row('Additional Shipping', 'Yes')

            # Proposition 65
            prop65_value = 'Y' if str(record.get('PartProposition65', '')).strip().lower() == 'yes' else 'N'
            add_attribute_row('Prop 65', prop65_value)

            processed_count += 1

        return {'processed_count': processed_count}

    def _populate_interchange(self, wb, product_data: List[Dict],
                              interchange_data: Dict[str, List[Dict]]) -> Dict[str, Any]:
        """Populate Interchange sheet"""
        ws = wb['Interchange']
        row_num = 1
        processed_count = 0

        for record in product_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue

            part_number = record.get('AS400_NumberStripped', '')
            part_type = record.get('SDC_PartType', '')

            # Get interchange records for this part
            interchange_records = interchange_data.get(part_number, [])

            # Filter by allowed interchange codes
            filtered_records = [
                r for r in interchange_records
                if r.get('ICPCD') in self.config.interchange_codes
            ]

            if not filtered_records:
                continue

            for interchange_record in filtered_records:
                row_num += 1
                try:
                    ws[f"A{row_num}"] = 'Default'
                    ws[f"B{row_num}"] = part_type
                    ws[f"C{row_num}"] = part_number
                    ws[f"D{row_num}"] = interchange_record.get('ICPNO', '')
                    ws[f"E{row_num}"] = 'GGMQ' if record.get('PartBrand') == 'RT Off-Road' else 'BKMN'

                    processed_count += 1

                except Exception as e:
                    logger.error(f"Error populating Interchange row {row_num}: {e}")
                    continue

        return {'processed_count': processed_count}