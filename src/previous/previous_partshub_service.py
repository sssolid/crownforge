"""
Partshub Template Service
Combines populated template with additional data sources to create final template
"""

import logging
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional
import unicodedata
import re
from openpyxl import load_workbook

from .utils import performance_monitor, TextProcessor

logger = logging.getLogger(__name__)


class PartshubService:
    """Service for generating final Partshub template"""

    def __init__(self, config):
        self.config = config

    def sanitize_string(self, value: Any) -> str:
        """Sanitize string for Excel output"""
        if value and isinstance(value, str):
            # Normalize Unicode to decompose formatting marks
            value = unicodedata.normalize('NFKD', value)
            # Remove all control characters, invisible characters, and strip whitespace
            value = re.sub(r'[\x00-\x1F\x7F-\x9F\u2000-\u200F\u2028-\u202F\u205F-\u206F\xa0]', '', value)
            # Remove Trademark symbol (™) and Registered Trademark symbol (®)
            value = value.replace('™', '').replace('®', '')
            # Remove other restricted characters
            value = value.replace('*', '').replace('\r', '').replace('\n', '').strip()
            return value
        return value if value is not None else ''

    def sanitize_workbook(self, wb) -> list:
        """Sanitize all cells in a workbook and report any changes"""
        changes = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        sanitized_value = self.sanitize_string(original_value)
                        if original_value != sanitized_value:
                            changes.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'original': original_value,
                                'sanitized': sanitized_value
                            })
                            cell.value = sanitized_value
        return changes

    def generate_final_template(self, popularity_file: str, applications_file: str,
                                populated_template: str, full_pies_file: str,
                                output_file: str) -> Dict[str, Any]:
        """
        Generate final Partshub template by combining all data sources

        Args:
            popularity_file: Path to popularity codes CSV
            applications_file: Path to application data Excel
            populated_template: Path to populated SDC template
            full_pies_file: Path to full PIES data Excel
            output_file: Path for final template output

        Returns:
            Dictionary with generation results
        """
        logger.info("Generating final Partshub template")

        with performance_monitor("Partshub Template Generation") as monitor:
            # Step 1: Load the populated template
            wb = load_workbook(populated_template)

            # Get existing sheets
            pies_extended_info = wb['PiesExtendedInfo']
            pies_descriptions = wb['PiesDescriptions']

            # Step 2: Load Popularity Codes and create mapping
            popularity_mapping = self._load_popularity_codes(popularity_file)
            monitor.increment_processed(len(popularity_mapping))

            # Step 3: Load Jeep Applications and create mapping
            jeep_mapping = self._load_jeep_applications(applications_file)
            monitor.increment_processed(len(jeep_mapping))

            # Step 4: Update PiesExtendedInfo with Popularity Codes
            popularity_updates = self._update_popularity_codes(pies_extended_info, popularity_mapping)

            # Step 5: Update PiesDescriptions with Jeep Applications
            application_updates = self._update_application_summaries(pies_descriptions, jeep_mapping)

            # Step 6: Populate PiesDigitalAsset sheet from full PIES data
            digital_asset_updates = self._populate_digital_asset_sheet(wb, full_pies_file)

            # Step 7: Sanitize all cells in the workbook
            changes = self.sanitize_workbook(wb)

            # Step 8: Save the final template
            Path(output_file).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)

            summary = {
                'input_files': {
                    'popularity_file': popularity_file,
                    'applications_file': applications_file,
                    'populated_template': populated_template,
                    'full_pies_file': full_pies_file
                },
                'output_file': output_file,
                'updates': {
                    'popularity_codes': popularity_updates,
                    'application_summaries': application_updates,
                    'digital_assets': digital_asset_updates
                },
                'sanitization_changes': len(changes),
                'total_changes': popularity_updates + application_updates + digital_asset_updates + len(changes)
            }

            logger.info(f"Final Partshub template generated: {summary}")
            return summary

    def _load_popularity_codes(self, popularity_file: str) -> Dict[str, str]:
        """Load popularity codes from CSV file"""
        try:
            popularity_df = pd.read_csv(popularity_file)
            return dict(zip(popularity_df['Part Number'], popularity_df['Popularity Code']))
        except FileNotFoundError:
            logger.warning(f"Popularity codes file not found: {popularity_file}")
            return {}
        except Exception as e:
            logger.error(f"Error loading popularity codes: {e}")
            return {}

    def _load_jeep_applications(self, applications_file: str) -> Dict[str, str]:
        """Load Jeep applications from Excel file"""
        try:
            jeep_df = pd.read_excel(applications_file, sheet_name='Jeep Applications')

            # Sanitize applications data
            jeep_mapping = {}
            for part_number, application in zip(jeep_df['PartNumber'], jeep_df['Application']):
                sanitized_application = self.sanitize_string(application)
                jeep_mapping[part_number] = sanitized_application

            return jeep_mapping
        except FileNotFoundError:
            logger.warning(f"Applications file not found: {applications_file}")
            return {}
        except Exception as e:
            logger.error(f"Error loading Jeep applications: {e}")
            return {}

    def _update_popularity_codes(self, sheet, popularity_mapping: Dict[str, str]) -> int:
        """Update PiesExtendedInfo sheet with popularity codes"""
        updates = 0

        for row in range(3, sheet.max_row + 1):
            part_number = sheet[f'C{row}'].value  # Column C has the Part Number

            if part_number and part_number in popularity_mapping:
                sheet[f'AP{row}'].value = popularity_mapping[part_number]  # Column AP for Popularity Code
                updates += 1

        logger.info(f"Updated {updates} popularity codes in PiesExtendedInfo")
        return updates

    def _update_application_summaries(self, sheet, jeep_mapping: Dict[str, str]) -> int:
        """Update PiesDescriptions sheet with Jeep applications"""
        updates = 0

        for row in range(3, sheet.max_row + 1):
            part_number = sheet[f'C{row}'].value  # Column C has the Part Number

            if part_number and part_number in jeep_mapping:
                sheet[f'N{row}'].value = jeep_mapping[part_number]  # Column N for Application Summary
                updates += 1

        logger.info(f"Updated {updates} application summaries in PiesDescriptions")
        return updates

    def _populate_digital_asset_sheet(self, wb, full_pies_file: str) -> int:
        """Populate PiesDigitalAsset sheet from full PIES data"""
        if not Path(full_pies_file).exists():
            logger.warning(f"Full PIES file not found: {full_pies_file}")
            return 0

        try:
            # Load the full PIES workbook
            full_wb = load_workbook(full_pies_file)

            if 'PiesDigitalAsset' not in full_wb.sheetnames:
                logger.warning("PiesDigitalAsset sheet not found in full PIES file")
                return 0

            source_sheet = full_wb['PiesDigitalAsset']

            # Extract matching part numbers from the final template
            matching_part_numbers = set()
            extended_info_sheet = wb['PiesExtendedInfo']

            for row in range(3, extended_info_sheet.max_row + 1):
                part_number = extended_info_sheet[f'C{row}'].value  # Column C holds the Part Number
                if part_number:
                    matching_part_numbers.add(part_number)

            # Get the destination sheet from the final template
            if 'PiesDigitalAsset' not in wb.sheetnames:
                dest_sheet = wb.create_sheet('PiesDigitalAsset')
                # Copy header row if it exists
                if source_sheet.max_row >= 1:
                    header_row = [cell.value for cell in source_sheet[1]]
                    dest_sheet.append(header_row)
            else:
                dest_sheet = wb['PiesDigitalAsset']

            # Copy matching rows
            added_rows = 0
            for row in source_sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 2:  # Ensure row has enough columns
                    part_number = row[2]  # Column C in source sheet (0-indexed: index 2)
                    if part_number in matching_part_numbers:
                        dest_sheet.append(row)
                        added_rows += 1

            logger.info(f"Added {added_rows} digital asset records")
            return added_rows

        except Exception as e:
            logger.error(f"Error populating digital asset sheet: {e}")
            return 0

    def validate_final_template(self, template_file: str) -> Dict[str, Any]:
        """Validate the final template for completeness and accuracy"""
        validation_results = {
            'file_exists': Path(template_file).exists(),
            'sheets_present': [],
            'missing_sheets': [],
            'data_summary': {},
            'issues': []
        }

        if not validation_results['file_exists']:
            validation_results['issues'].append(f"Template file not found: {template_file}")
            return validation_results

        try:
            wb = load_workbook(template_file)

            # Expected sheets
            expected_sheets = [
                'PiesItem', 'PiesDescriptions', 'PiesPrices',
                'PiesExtendedInfo', 'PiesPackages', 'PiesUserAttr',
                'Interchange', 'PiesDigitalAsset'
            ]

            for sheet_name in expected_sheets:
                if sheet_name in wb.sheetnames:
                    validation_results['sheets_present'].append(sheet_name)
                    sheet = wb[sheet_name]
                    validation_results['data_summary'][sheet_name] = {
                        'rows': sheet.max_row,
                        'columns': sheet.max_column
                    }
                else:
                    validation_results['missing_sheets'].append(sheet_name)

            # Check for data completeness
            if 'PiesExtendedInfo' in wb.sheetnames:
                sheet = wb['PiesExtendedInfo']
                empty_popularity_codes = 0

                for row in range(3, sheet.max_row + 1):
                    if not sheet[f'AP{row}'].value:  # Popularity code column
                        empty_popularity_codes += 1

                if empty_popularity_codes > 0:
                    validation_results['issues'].append(
                        f"{empty_popularity_codes} records missing popularity codes"
                    )

            if 'PiesDescriptions' in wb.sheetnames:
                sheet = wb['PiesDescriptions']
                empty_applications = 0

                for row in range(3, sheet.max_row + 1):
                    if not sheet[f'N{row}'].value:  # Application summary column
                        empty_applications += 1

                if empty_applications > 0:
                    validation_results['issues'].append(
                        f"{empty_applications} records missing application summaries"
                    )

            validation_results['is_valid'] = len(validation_results['issues']) == 0

        except Exception as e:
            validation_results['issues'].append(f"Error validating template: {e}")
            validation_results['is_valid'] = False

        return validation_results