"""
Workflow Service
Orchestrates the complete data processing workflow
"""

import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from pathlib import Path

from .application_parser import ApplicationParserService
from .popularity_service import PopularityService
from .sdc_service import SDCService
from .validators.upc_validator import UPCValidator
from .database import FilemakerService, IseriesService
from .output_service import ExcelOutputService
from .utils import performance_monitor, ConfigManager

logger = logging.getLogger(__name__)

@dataclass
class WorkflowConfig:
    """Workflow configuration"""
    enabled_steps: List[str]
    step_dependencies: Dict[str, List[str]]

@dataclass
class WorkflowResult:
    """Result of workflow execution"""
    success: bool
    completed_steps: List[str]
    failed_steps: List[str]
    results: Dict[str, Any]
    errors: List[str]

class WorkflowService:
    """Main workflow orchestrator"""

    def __init__(self, config_manager: ConfigManager):
        self.config = config_manager
        self.workflow_config = WorkflowConfig(
            enabled_steps=config_manager.get('workflow.enabled_steps', []),
            step_dependencies=config_manager.get('workflow.step_dependencies', {})
        )

        # Initialize services
        self._initialize_services()

        # Track workflow state
        self.completed_steps = []
        self.failed_steps = []
        self.step_results = {}

    def _initialize_services(self):
        """Initialize all required services"""

        # Database services
        filemaker_config = self.config.get('database.filemaker')
        iseries_config = self.config.get('database.iseries')
        query_templates_dir = self.config.get('files.query_templates_dir')

        self.filemaker_service = FilemakerService(
            filemaker_config, query_templates_dir
        )

        self.iseries_service = IseriesService(
            iseries_config, query_templates_dir
        )

        # Workflow services
        popularity_config = self.config.get('popularity_codes')
        self.popularity_service = PopularityService(
            self.iseries_service, self.filemaker_service, popularity_config
        )

        sdc_config = self.config.get('sdc_template')
        self.sdc_service = SDCService(self.filemaker_service, sdc_config)

        # Validation services
        upc_config = self.config.get('validation.upc_validation')
        self.upc_validator = UPCValidator(upc_config)

        measurement_config = self.config.get('validation.measurement_validation')
        from .validators.measurement_validator import MeasurementValidator, MeasurementValidationConfig
        self.measurement_validator = MeasurementValidator(measurement_config)

        # Application parser (refactored)
        from .application_parser import create_application_parser_service
        app_config = self.config.get('validation')  # Use validation config section
        self.application_parser_service = create_application_parser_service(
            self.filemaker_service, app_config
        )

    def execute_workflow(self) -> WorkflowResult:
        """Execute the complete workflow"""
        logger.info("Starting complete data processing workflow")

        with performance_monitor("Complete Workflow") as monitor:
            errors = []

            # Execute steps in dependency order
            for step in self.workflow_config.enabled_steps:
                if not self._check_dependencies(step):
                    logger.warning(f"Skipping step '{step}' - dependencies not met")
                    continue

                try:
                    logger.info(f"Executing workflow step: {step}")

                    if step == "applications":
                        result = self._execute_applications_step()
                    elif step == "popularity_codes":
                        result = self._execute_popularity_codes_step()
                    elif step == "sdc_template":
                        result = self._execute_sdc_template_step()
                    elif step == "partshub_template":
                        result = self._execute_partshub_template_step()
                    elif step == "validations":
                        result = self._execute_validations_step()
                    else:
                        logger.warning(f"Unknown workflow step: {step}")
                        continue

                    self.step_results[step] = result
                    self.completed_steps.append(step)
                    monitor.increment_processed(1)

                    logger.info(f"Step '{step}' completed successfully")

                except Exception as e:
                    error_msg = f"Step '{step}' failed: {e}"
                    logger.error(error_msg)
                    errors.append(error_msg)
                    self.failed_steps.append(step)
                    monitor.increment_errors(1)

                    if not self.config.get('error_handling.continue_on_error', True):
                        break

            success = len(self.failed_steps) == 0

            return WorkflowResult(
                success=success,
                completed_steps=self.completed_steps,
                failed_steps=self.failed_steps,
                results=self.step_results,
                errors=errors
            )

    def _check_dependencies(self, step: str) -> bool:
        """Check if step dependencies are satisfied"""
        dependencies = self.workflow_config.step_dependencies.get(step, [])

        for dep in dependencies:
            if dep not in self.completed_steps:
                return False

        return True

    def _execute_applications_step(self) -> Dict[str, Any]:
        """Execute application processing step"""
        if not self.config.get('processing.features.extract_engines', True):
            logger.info("Application processing disabled in configuration")
            return {"status": "skipped", "reason": "disabled in config"}

        # Execute application parser service
        result = self.application_parser_service.process_all()

        if result.success:
            # Generate Excel output
            output_file = self.config.get('files.application_data')
            output_service = ExcelOutputService(output_file)

            if output_service.generate_excel_report(self.application_parser_service.results):
                return {
                    "status": "success",
                    "output_file": output_file,
                    "processed_count": len(self.application_parser_service.results.get('correct_applications', []))
                }
            else:
                raise Exception("Failed to generate Excel report")
        else:
            raise Exception(f"Application processing failed: {result.errors}")

    def _execute_popularity_codes_step(self) -> Dict[str, Any]:
        """Execute popularity codes generation step"""
        if not self.config.get('processing.features.generate_popularity_codes', True):
            logger.info("Popularity codes generation disabled in configuration")
            return {"status": "skipped", "reason": "disabled in config"}

        output_file = self.config.get('files.popularity_codes')
        branch = self.config.get('popularity_codes.default_branch')
        brand = self.config.get('popularity_codes.default_brand')
        start_date = self.config.get('popularity_codes.default_start_date')

        result = self.popularity_service.generate_popularity_codes(
            output_file=output_file,
            branch=branch,
            brand=brand,
            start_date=start_date
        )

        return {
            "status": "success",
            "output_file": output_file,
            **result
        }

    def _execute_sdc_template_step(self) -> Dict[str, Any]:
        """Execute SDC template population step"""
        if not self.config.get('processing.features.populate_sdc_template', True):
            logger.info("SDC template population disabled in configuration")
            return {"status": "skipped", "reason": "disabled in config"}

        template_file = self.config.get('files.sdc_blank_template')
        output_file = self.config.get('files.sdc_populated_template')
        missing_parts_file = self.config.get('files.missing_parts_list')

        # Load popularity mapping if available
        popularity_mapping = None
        popularity_file = self.config.get('files.popularity_codes')
        if Path(popularity_file).exists():
            popularity_mapping = self.popularity_service.get_popularity_mapping(popularity_file)

        result = self.sdc_service.populate_sdc_template(
            template_file=template_file,
            output_file=output_file,
            missing_parts_file=missing_parts_file,
            popularity_mapping=popularity_mapping
        )

        return {
            "status": "success",
            **result
        }

    def _execute_partshub_template_step(self) -> Dict[str, Any]:
        """Execute Partshub template generation step"""
        if not self.config.get('processing.features.generate_partshub_template', True):
            logger.info("Partshub template generation disabled in configuration")
            return {"status": "skipped", "reason": "disabled in config"}

        from .partshub_service import PartshubService

        # Initialize Partshub service
        partshub_service = PartshubService(self.config)

        # Get file paths from config
        popularity_file = self.config.get('files.popularity_codes')
        applications_file = self.config.get('files.application_data')
        populated_template = self.config.get('files.sdc_populated_template')
        full_pies_file = self.config.get('files.sdc_full_pies')
        output_file = self.config.get('files.sdc_final_template')

        # Generate final template
        result = partshub_service.generate_final_template(
            popularity_file=popularity_file,
            applications_file=applications_file,
            populated_template=populated_template,
            full_pies_file=full_pies_file,
            output_file=output_file
        )

        # Validate the final template
        validation_result = partshub_service.validate_final_template(output_file)
        result['validation'] = validation_result

        return {
            "status": "success",
            **result
        }

    def _execute_validations_step(self) -> Dict[str, Any]:
        """Execute data validation step"""
        validation_results = {}

        # UPC Validation
        if self.config.get('validation.upc_validation.enabled', True):
            upc_data = self.filemaker_service.get_upc_validation_data()
            upc_results = self.upc_validator.validate_dataset(upc_data)
            validation_results['upc_validation'] = upc_results

            # Generate UPC validation report
            upc_report_file = self.config.get('files.upc_validation_report')
            self._generate_upc_validation_report(upc_results, upc_report_file)

        # Measurement Validation
        if self.config.get('validation.measurement_validation.enabled', True):
            measurement_results = self._validate_measurements()
            validation_results['measurement_validation'] = measurement_results

        # Kit Components Analysis
        if self.config.get('processing.features.generate_kit_components', True):
            kit_results = self._analyze_kit_components()
            validation_results['kit_components'] = kit_results

        # Cost Discrepancies
        if self.config.get('processing.features.check_cost_discrepancies', True):
            cost_results = self._analyze_cost_discrepancies()
            validation_results['cost_discrepancies'] = cost_results

        return {
            "status": "success",
            "validation_results": validation_results
        }

    def _generate_upc_validation_report(self, upc_results: Dict, output_file: str):
        """Generate UPC validation Excel report"""
        from openpyxl import Workbook

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary = upc_results['summary']
        summary_data = [
            ['Metric', 'Count'],
            ['Total Records', summary['total_records']],
            ['Valid UPCs', summary['valid_upcs']],
            ['Invalid UPCs', summary['invalid_upcs']],
            ['Empty UPCs', summary['empty_upcs']],
            ['Duplicate UPCs', summary['duplicate_count']],
            ['Validation Rate (%)', f"{summary['validation_rate']:.2f}"]
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)

        # Invalid UPCs sheet
        if upc_results['invalid_records']:
            ws_invalid = wb.create_sheet("Invalid UPCs")
            headers = ['Part Number', 'UPC', 'Error Message', 'Suggested Check Digit']

            for col_idx, header in enumerate(headers, 1):
                ws_invalid.cell(row=1, column=col_idx, value=header)

            for row_idx, record in enumerate(upc_results['invalid_records'], 2):
                ws_invalid.cell(row=row_idx, column=1, value=record['PartNumber'])
                ws_invalid.cell(row=row_idx, column=2, value=record['UPC'])
                ws_invalid.cell(row=row_idx, column=3, value=record['ErrorMessage'])
                ws_invalid.cell(row=row_idx, column=4, value=record['SuggestedCheckDigit'])

        # Duplicate UPCs sheet
        if upc_results['duplicate_upcs']:
            ws_duplicates = wb.create_sheet("Duplicate UPCs")
            headers = ['UPC', 'Count', 'Part Numbers']

            for col_idx, header in enumerate(headers, 1):
                ws_duplicates.cell(row=1, column=col_idx, value=header)

            for row_idx, record in enumerate(upc_results['duplicate_upcs'], 2):
                ws_duplicates.cell(row=row_idx, column=1, value=record['UPC'])
                ws_duplicates.cell(row=row_idx, column=2, value=record['Count'])
                ws_duplicates.cell(row=row_idx, column=3, value=record['PartNumbers'])

        # Save workbook
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        logger.info(f"UPC validation report saved to {output_file}")

    def _analyze_kit_components(self) -> Dict[str, Any]:
        """Analyze kit components and hierarchy"""
        try:
            kit_data = self.iseries_service.get_kit_components_hierarchy()

            # Process and analyze the data
            analysis = {
                'total_components': len(kit_data),
                'unique_assemblies': len(set(r['Assembly'] for r in kit_data)),
                'cost_discrepancies': [r for r in kit_data if r.get('CostDiscrepancy', 0) != 0],
                'data': kit_data
            }

            # Generate report
            report_file = self.config.get('files.kit_components_report')
            self._generate_kit_components_report(analysis, report_file)

            return analysis

        except Exception as e:
            logger.error(f"Kit components analysis failed: {e}")
            return {'error': str(e)}

    def _analyze_cost_discrepancies(self) -> Dict[str, Any]:
        """Analyze cost discrepancies between systems"""
        try:
            cost_data = self.iseries_service.validate_cost_discrepancies()

            analysis = {
                'total_discrepancies': len(cost_data),
                'data': cost_data
            }

            # Generate report
            report_file = self.config.get('files.cost_discrepancy_report')
            self._generate_cost_discrepancy_report(analysis, report_file)

            return analysis

        except Exception as e:
            logger.error(f"Cost discrepancy analysis failed: {e}")
            return {'error': str(e)}

    def _validate_measurements(self) -> Dict[str, Any]:
        """Validate measurements between Filemaker and Iseries"""
        try:
            # Get measurement data from both systems
            fm_data = self.filemaker_service.validate_measurements()
            is_data = self.iseries_service.get_measurement_data()

            # Validate each dataset individually
            fm_validation = self.measurement_validator.validate_dataset(fm_data, "Filemaker")
            is_validation = self.measurement_validator.validate_dataset(is_data, "Iseries")

            # Compare measurements between systems
            discrepancies = self.measurement_validator.compare_measurements(fm_data, is_data)
            discrepancy_report = self.measurement_validator.generate_discrepancy_report(discrepancies)

            analysis = {
                'filemaker_validation': fm_validation,
                'iseries_validation': is_validation,
                'discrepancies': discrepancies,
                'discrepancy_report': discrepancy_report,
                'total_discrepancies': len(discrepancies),
                'significant_discrepancies': len([d for d in discrepancies if d.is_significant])
            }

            # Generate measurement validation report
            self._generate_measurement_validation_report(analysis)

            return analysis

        except Exception as e:
            logger.error(f"Measurement validation failed: {e}")
            return {'error': str(e)}

    def _generate_measurement_validation_report(self, analysis: Dict):
        """Generate measurement validation Excel report"""
        from openpyxl import Workbook

        output_file = self.config.get("files.measurement_validation_report")
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_data = [
            ['Metric', 'Filemaker', 'Iseries'],
            ['Total Records', analysis['filemaker_validation']['total_records'],
             analysis['iseries_validation']['total_records']],
            ['Valid Records', analysis['filemaker_validation']['valid_records'],
             analysis['iseries_validation']['valid_records']],
            ['Invalid Records', analysis['filemaker_validation']['invalid_records'],
             analysis['iseries_validation']['invalid_records']],
            ['Total Discrepancies', analysis['total_discrepancies'], ''],
            ['Significant Discrepancies', analysis['significant_discrepancies'], '']
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)

        # Discrepancies sheet
        if analysis['discrepancies']:
            ws_disc = wb.create_sheet("Discrepancies")
            headers = ['Part Number', 'Field', 'Filemaker Value', 'Iseries Value',
                      'Difference', 'Percentage Diff', 'Significant']

            for col_idx, header in enumerate(headers, 1):
                ws_disc.cell(row=1, column=col_idx, value=header)

            for row_idx, disc in enumerate(analysis['discrepancies'], 2):
                ws_disc.cell(row=row_idx, column=1, value=disc.part_number)
                ws_disc.cell(row=row_idx, column=2, value=disc.field)
                ws_disc.cell(row=row_idx, column=3, value=disc.filemaker_value)
                ws_disc.cell(row=row_idx, column=4, value=disc.iseries_value)
                ws_disc.cell(row=row_idx, column=5, value=disc.difference)
                ws_disc.cell(row=row_idx, column=6, value=disc.percentage_diff)
                ws_disc.cell(row=row_idx, column=7, value=disc.is_significant)

        # Save workbook
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        logger.info(f"Measurement validation report saved to {output_file}")

    def _generate_kit_components_report(self, analysis: Dict, output_file: str):
        """Generate kit components Excel report"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime

        try:
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Summary data
            summary_data = [
                ['Metric', 'Count'],
                ['Total Components', analysis.get('total_components', 0)],
                ['Unique Assemblies', analysis.get('unique_assemblies', 0)],
                ['Cost Discrepancies Found', len(analysis.get('cost_discrepancies', []))],
                ['Report Generated', datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]

            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in ws_summary.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                ws_summary.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

            # Kit Components Data sheet
            if analysis.get('data'):
                ws_data = wb.create_sheet("Kit Components")

                # Headers
                headers = ['Assembly', 'Component', 'Quantity', 'Level', 'Cost From INSMFH',
                           'Latest Component Cost', 'Cost Discrepancy']

                for col_idx, header in enumerate(headers, 1):
                    cell = ws_data.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Data rows
                for row_idx, record in enumerate(analysis['data'], 2):
                    ws_data.cell(row=row_idx, column=1, value=record.get('Assembly', ''))
                    ws_data.cell(row=row_idx, column=2, value=record.get('Component', ''))
                    ws_data.cell(row=row_idx, column=3, value=record.get('Quantity', ''))
                    ws_data.cell(row=row_idx, column=4, value=record.get('Level', ''))
                    ws_data.cell(row=row_idx, column=5, value=record.get('CostFromINSMFH', ''))
                    ws_data.cell(row=row_idx, column=6, value=record.get('LatestComponentCost', ''))

                    # Highlight discrepancies
                    discrepancy = record.get('CostDiscrepancy', 0)
                    discrepancy_cell = ws_data.cell(row=row_idx, column=7, value=discrepancy)
                    if discrepancy != 0:
                        discrepancy_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                # Auto-adjust column widths
                for column in ws_data.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value)
                    ws_data.column_dimensions[column[0].column_letter].width = min(max_length + 2, 30)

            # Cost Discrepancies sheet (if any found)
            cost_discrepancies = analysis.get('cost_discrepancies', [])
            if cost_discrepancies:
                ws_disc = wb.create_sheet("Cost Discrepancies")

                headers = ['Assembly', 'Component', 'Quantity', 'INSMFH Cost',
                           'Component Cost', 'Discrepancy Amount', 'Impact (Qty × Disc)']

                for col_idx, header in enumerate(headers, 1):
                    cell = ws_disc.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                for row_idx, record in enumerate(cost_discrepancies, 2):
                    quantity = record.get('Quantity', 1)
                    discrepancy = record.get('CostDiscrepancy', 0)
                    impact = quantity * discrepancy if quantity and discrepancy else 0

                    ws_disc.cell(row=row_idx, column=1, value=record.get('Assembly', ''))
                    ws_disc.cell(row=row_idx, column=2, value=record.get('Component', ''))
                    ws_disc.cell(row=row_idx, column=3, value=quantity)
                    ws_disc.cell(row=row_idx, column=4, value=record.get('CostFromINSMFH', ''))
                    ws_disc.cell(row=row_idx, column=5, value=record.get('LatestComponentCost', ''))
                    ws_disc.cell(row=row_idx, column=6, value=discrepancy)
                    impact_cell = ws_disc.cell(row=row_idx, column=7, value=impact)

                    # Highlight significant impacts
                    if abs(impact) > 10:  # Highlight impacts over $10
                        impact_cell.fill = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")

                # Auto-adjust column widths
                for column in ws_disc.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value)
                    ws_disc.column_dimensions[column[0].column_letter].width = min(max_length + 2, 25)

            # Save workbook
            Path(output_file).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)
            logger.info(f"Kit components report generated: {output_file}")

        except Exception as e:
            logger.error(f"Failed to generate kit components report: {e}")

    def _generate_cost_discrepancy_report(self, analysis: Dict, output_file: str):
        """Generate cost discrepancy Excel report"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        from collections import defaultdict

        try:
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            # Summary statistics
            total_discrepancies = analysis.get('total_discrepancies', 0)
            data = analysis.get('data', [])

            # Calculate statistics
            if data:
                total_cost_diff = sum(abs(record.get('CostDiscrepancy', 0)) for record in data)
                avg_discrepancy = total_cost_diff / len(data) if data else 0
                max_discrepancy = max((abs(record.get('CostDiscrepancy', 0)) for record in data), default=0)
                significant_discrepancies = len([r for r in data if abs(r.get('CostDiscrepancy', 0)) > 5])
            else:
                total_cost_diff = avg_discrepancy = max_discrepancy = significant_discrepancies = 0

            summary_data = [
                ['Cost Discrepancy Analysis Summary', ''],
                ['', ''],
                ['Metric', 'Value'],
                ['Total Discrepancies Found', total_discrepancies],
                ['Significant Discrepancies (>$5)', significant_discrepancies],
                ['Total Cost Difference', f"${total_cost_diff:.2f}"],
                ['Average Discrepancy', f"${avg_discrepancy:.2f}"],
                ['Maximum Discrepancy', f"${max_discrepancy:.2f}"],
                ['Report Generated', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ['', ''],
                ['Significance Levels:', ''],
                ['🟢 Minor (<$1)', 'Review when convenient'],
                ['🟡 Moderate ($1-$5)', 'Review within 30 days'],
                ['🔴 Significant (>$5)', 'Review immediately']
            ]

            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)

                    if row_idx == 1:  # Title
                        cell.font = Font(bold=True, size=14)
                        cell.fill = header_fill
                    elif row_idx == 3:  # Headers
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    elif 'Significant' in str(value) and significant_discrepancies > 0:
                        cell.fill = error_fill
                    elif 'Total Cost' in str(value) and total_cost_diff > 100:
                        cell.fill = warning_fill

            # Merge title cells
            ws_summary.merge_cells('A1:B1')

            # Auto-adjust column widths
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20

            # Detailed Discrepancies sheet
            if data:
                ws_details = wb.create_sheet("Detailed Discrepancies")

                headers = ['Assembly', 'Component', 'Quantity', 'Level', 'INSMFH Cost',
                           'Component Cost', 'Discrepancy', 'Total Impact', 'Significance']

                for col_idx, header in enumerate(headers, 1):
                    cell = ws_details.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Sort data by discrepancy amount (highest first)
                sorted_data = sorted(data, key=lambda x: abs(x.get('CostDiscrepancy', 0)), reverse=True)

                for row_idx, record in enumerate(sorted_data, 2):
                    quantity = record.get('Quantity', 1) or 1
                    discrepancy = record.get('CostDiscrepancy', 0) or 0
                    total_impact = quantity * abs(discrepancy)

                    # Determine significance
                    abs_disc = abs(discrepancy)
                    if abs_disc > 5:
                        significance = "🔴 Critical"
                        fill_color = error_fill
                    elif abs_disc > 1:
                        significance = "🟡 Moderate"
                        fill_color = warning_fill
                    else:
                        significance = "🟢 Minor"
                        fill_color = None

                    # Populate row
                    cells = [
                        ws_details.cell(row=row_idx, column=1, value=record.get('Assembly', '')),
                        ws_details.cell(row=row_idx, column=2, value=record.get('Component', '')),
                        ws_details.cell(row=row_idx, column=3, value=quantity),
                        ws_details.cell(row=row_idx, column=4, value=record.get('Level', '')),
                        ws_details.cell(row=row_idx, column=5, value=f"${record.get('CostFromINSMFH', 0):.2f}"),
                        ws_details.cell(row=row_idx, column=6, value=f"${record.get('LatestComponentCost', 0):.2f}"),
                        ws_details.cell(row=row_idx, column=7, value=f"${discrepancy:.2f}"),
                        ws_details.cell(row=row_idx, column=8, value=f"${total_impact:.2f}"),
                        ws_details.cell(row=row_idx, column=9, value=significance)
                    ]

                    # Apply significance coloring
                    if fill_color:
                        for cell in cells:
                            cell.fill = fill_color

                # Auto-adjust column widths
                for col_idx in range(1, 10):
                    max_length = max(
                        len(str(ws_details.cell(row=row, column=col_idx).value))
                        for row in range(1, ws_details.max_row + 1)
                    )
                    ws_details.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 25)

            # Assembly Summary sheet
            if data:
                ws_assembly = wb.create_sheet("Assembly Summary")

                # Group by assembly
                assembly_stats = defaultdict(lambda: {
                    'component_count': 0,
                    'total_discrepancy': 0,
                    'max_discrepancy': 0,
                    'components': []
                })

                for record in data:
                    assembly = record.get('Assembly', 'Unknown')
                    discrepancy = abs(record.get('CostDiscrepancy', 0))

                    assembly_stats[assembly]['component_count'] += 1
                    assembly_stats[assembly]['total_discrepancy'] += discrepancy
                    assembly_stats[assembly]['max_discrepancy'] = max(
                        assembly_stats[assembly]['max_discrepancy'], discrepancy
                    )
                    assembly_stats[assembly]['components'].append(record.get('Component', ''))

                headers = ['Assembly', 'Component Count', 'Total Discrepancy',
                           'Max Single Discrepancy', 'Avg Discrepancy', 'Priority']

                for col_idx, header in enumerate(headers, 1):
                    cell = ws_assembly.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                # Sort assemblies by total discrepancy
                sorted_assemblies = sorted(
                    assembly_stats.items(),
                    key=lambda x: x[1]['total_discrepancy'],
                    reverse=True
                )

                for row_idx, (assembly, stats) in enumerate(sorted_assemblies, 2):
                    avg_discrepancy = stats['total_discrepancy'] / stats['component_count']

                    # Determine priority
                    if stats['total_discrepancy'] > 50 or stats['max_discrepancy'] > 10:
                        priority = "🔴 High"
                        priority_fill = error_fill
                    elif stats['total_discrepancy'] > 20 or stats['max_discrepancy'] > 5:
                        priority = "🟡 Medium"
                        priority_fill = warning_fill
                    else:
                        priority = "🟢 Low"
                        priority_fill = None

                    cells = [
                        ws_assembly.cell(row=row_idx, column=1, value=assembly),
                        ws_assembly.cell(row=row_idx, column=2, value=stats['component_count']),
                        ws_assembly.cell(row=row_idx, column=3, value=f"${stats['total_discrepancy']:.2f}"),
                        ws_assembly.cell(row=row_idx, column=4, value=f"${stats['max_discrepancy']:.2f}"),
                        ws_assembly.cell(row=row_idx, column=5, value=f"${avg_discrepancy:.2f}"),
                        ws_assembly.cell(row=row_idx, column=6, value=priority)
                    ]

                    if priority_fill:
                        for cell in cells:
                            cell.fill = priority_fill

                # Auto-adjust column widths
                for col_idx in range(1, 7):
                    max_length = max(
                        len(str(ws_assembly.cell(row=row, column=col_idx).value))
                        for row in range(1, ws_assembly.max_row + 1)
                    )
                    ws_assembly.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 25)

            # Save workbook
            Path(output_file).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)
            logger.info(f"Cost discrepancy report generated: {output_file} with {total_discrepancies} discrepancies")

        except Exception as e:
            logger.error(f"Failed to generate cost discrepancy report: {e}")

    def get_workflow_status(self) -> Dict[str, Any]:
        """Get current workflow status"""
        return {
            'enabled_steps': self.workflow_config.enabled_steps,
            'completed_steps': self.completed_steps,
            'failed_steps': self.failed_steps,
            'pending_steps': [
                step for step in self.workflow_config.enabled_steps
                if step not in self.completed_steps and step not in self.failed_steps
            ]
        }