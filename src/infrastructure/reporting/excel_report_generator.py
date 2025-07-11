# src/infrastructure/reporting/excel_report_generator.py
"""
Excel report generator with advanced formatting and multiple sheet support.
"""

import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from dataclasses import dataclass
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from ...domain.interfaces import ReportGenerator
from ...domain.models import ProcessingResult

logger = logging.getLogger(__name__)


@dataclass
class ExcelReportConfig:
    """Excel report configuration."""
    include_formatting: bool = True
    auto_filter: bool = True
    freeze_headers: bool = True
    max_column_width: int = 50
    add_summary_sheet: bool = True
    add_charts: bool = False
    protect_sheets: bool = False


@dataclass
class SheetDefinition:
    """Definition for an Excel sheet."""
    name: str
    data_key: str
    description: str
    sort_columns: Optional[List[str]] = None
    freeze_row: int = 1
    conditional_formatting: Optional[Dict[str, Any]] = None


class ExcelReportGenerator(ReportGenerator):
    """Advanced Excel report generator."""

    def __init__(self, config: ExcelReportConfig):
        self.config = config
        self.predefined_sheets = self._initialize_sheet_definitions()

    def generate_report(self, data: Dict[str, Any], output_path: str) -> ProcessingResult:
        """Generate comprehensive Excel report."""
        try:
            logger.info(f"Generating Excel report: {output_path}")

            # Ensure output directory exists
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Add summary sheet if requested
            if self.config.add_summary_sheet:
                self._add_summary_sheet(wb, data)

            # Process each data section
            sheets_created = 0
            for data_key, sheet_data in data.items():
                if self._should_create_sheet(data_key, sheet_data):
                    sheet_created = self._create_data_sheet(wb, data_key, sheet_data)
                    if sheet_created:
                        sheets_created += 1

            # Apply workbook-level formatting
            if self.config.include_formatting:
                self._apply_workbook_formatting(wb)

            # Save workbook
            wb.save(output_path)

            logger.info(f"Excel report generated with {sheets_created} data sheets: {output_path}")

            return ProcessingResult(
                success=True,
                items_processed=sheets_created,
                data={
                    'output_file': output_path,
                    'sheets_created': sheets_created,
                    'file_size_mb': self._get_file_size_mb(output_path)
                }
            )

        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return ProcessingResult(
                success=False,
                errors=[f"Excel report generation failed: {e}"]
            )

    def get_supported_formats(self) -> List[str]:
        """Get supported output formats."""
        return ['xlsx', 'xlsm']

    @staticmethod
    def _initialize_sheet_definitions() -> Dict[str, SheetDefinition]:
        """Initialize predefined sheet definitions."""
        return {
            'correct_applications': SheetDefinition(
                name='Correct Applications',
                data_key='correct_applications',
                description='Valid, correctly formatted vehicle applications',
                sort_columns=['PartNumber', 'YearStart']
            ),
            'incorrect_applications': SheetDefinition(
                name='Incorrect Applications',
                data_key='incorrect_applications',
                description='Applications with format or validation issues'
            ),
            'invalid_applications': SheetDefinition(
                name='Invalid Applications',
                data_key='invalid_lines',
                description='Applications that failed validation'
            ),
            'marketing_validation': SheetDefinition(
                name='Marketing Validation',
                data_key='marketing_validation',
                description='Marketing description validation results'
            ),
            'upc_validation': SheetDefinition(
                name='UPC Validation',
                data_key='upc_validation',
                description='UPC code validation results'
            ),
            'measurement_discrepancies': SheetDefinition(
                name='Measurement Discrepancies',
                data_key='measurement_discrepancies',
                description='Measurement differences between systems'
            ),
            'cost_discrepancies': SheetDefinition(
                name='Cost Discrepancies',
                data_key='cost_discrepancies',
                description='Cost differences between systems'
            )
        }

    @staticmethod
    def _should_create_sheet(_data_key: str, sheet_data: Any) -> bool:
        """Determine if sheet should be created for data."""
        if not sheet_data:
            return False

        if isinstance(sheet_data, list) and len(sheet_data) == 0:
            return False

        if isinstance(sheet_data, dict) and len(sheet_data) == 0:
            return False

        return True

    @staticmethod
    def _add_summary_sheet(workbook: Workbook, data: Dict[str, Any]) -> None:
        """Add summary sheet to workbook."""
        ws = workbook.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Report Summary"
        ws['A1'].font = Font(bold=True, size=16)

        # Generation info
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Data summary
        row = 5
        ws[f'A{row}'] = "Data Section"
        ws[f'B{row}'] = "Record Count"
        ws[f'C{row}'] = "Status"

        # Make header bold
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)

        row += 1
        for data_key, sheet_data in data.items():
            ws[f'A{row}'] = data_key.replace('_', ' ').title()

            # Count records
            if isinstance(sheet_data, list):
                count = len(sheet_data)
            elif isinstance(sheet_data, dict):
                count = len(sheet_data)
            else:
                count = 1 if sheet_data else 0

            ws[f'B{row}'] = count
            ws[f'C{row}'] = "✓" if count > 0 else "○"

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 2, 30)

    def _create_data_sheet(self, workbook: Workbook, data_key: str, sheet_data: Any) -> bool:
        """Create data sheet for specific data section."""
        try:
            # Get sheet definition
            sheet_def = self.predefined_sheets.get(data_key)
            if sheet_def:
                sheet_name = sheet_def.name
            else:
                sheet_name = data_key.replace('_', ' ').title()

            # Create sheet
            ws = workbook.create_sheet(sheet_name)

            # Convert data to DataFrame
            df = self._prepare_dataframe(sheet_data)
            if df.empty:
                return False

            # Sort if specified
            if sheet_def and sheet_def.sort_columns:
                sort_cols = [col for col in sheet_def.sort_columns if col in df.columns]
                if sort_cols:
                    df = df.sort_values(by=sort_cols)

            # Write data to sheet
            self._write_dataframe_to_sheet(ws, df)

            # Apply formatting
            if self.config.include_formatting:
                self._apply_sheet_formatting(ws, df, sheet_def)

            return True

        except Exception as e:
            logger.error(f"Failed to create sheet for {data_key}: {e}")
            return False

    @staticmethod
    def _prepare_dataframe(data: Any) -> pd.DataFrame:
        """Convert various data formats to DataFrame."""
        if isinstance(data, pd.DataFrame):
            return data
        elif isinstance(data, list):
            if not data:
                return pd.DataFrame()

            # Handle list of dictionaries
            if isinstance(data[0], dict):
                return pd.DataFrame(data)
            else:
                # Handle list of objects with __dict__
                dict_data = []
                for item in data:
                    if hasattr(item, '__dict__'):
                        dict_data.append(item.__dict__)
                    elif hasattr(item, '_asdict'):  # namedtuple
                        dict_data.append(item._asdict())
                    else:
                        dict_data.append({'value': item})
                return pd.DataFrame(dict_data)
        elif isinstance(data, dict):
            # Convert dict to DataFrame
            if all(isinstance(v, (list, tuple)) for v in data.values()):
                return pd.DataFrame(data)
            else:
                return pd.DataFrame([data])
        else:
            return pd.DataFrame([{'value': data}])

    def _write_dataframe_to_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """Write DataFrame to worksheet with proper formatting."""
        # Write headers
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column_name
            if self.config.include_formatting:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                try:
                    if value is None or (isinstance(value, float) and pd.isna(value)):
                        cell.value = None
                    elif isinstance(value, (int, float, str)):
                        cell.value = value
                    else:
                        cell.value = str(value)
                except Exception as e:
                    logger.warning(f"Failed to write cell [{row_idx}, {col_idx}]: {e}")
                    cell.value = str(value)

    def _apply_sheet_formatting(self, worksheet, df: pd.DataFrame, sheet_def: Optional[SheetDefinition]) -> None:
        """Apply formatting to sheet."""
        # Freeze headers
        if self.config.freeze_headers and sheet_def:
            freeze_row = sheet_def.freeze_row + 1
            worksheet.freeze_panes = f"A{freeze_row}"

        # Auto filter
        if self.config.auto_filter and df.shape[0] > 0:
            max_column = df.shape[1]
            max_col_letter = get_column_letter(max_column)
            worksheet.auto_filter.ref = f"A1:{max_col_letter}1"

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass

            adjusted_width = min(max_length + 2, self.config.max_column_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Apply conditional formatting if defined
        if sheet_def and sheet_def.conditional_formatting:
            self._apply_conditional_formatting(worksheet, df, sheet_def.conditional_formatting)

    @staticmethod
    def _apply_conditional_formatting(worksheet, df: pd.DataFrame, formatting_rules: Dict[str, Any]) -> None:
        """Apply conditional formatting rules."""
        for rule_name, rule_config in formatting_rules.items():
            try:
                if rule_config['type'] == 'color_scale':
                    # Apply color scale to specified columns
                    for col_name in rule_config.get('columns', []):
                        if col_name in df.columns:
                            col_idx = df.columns.get_loc(col_name) + 1
                            col_letter = get_column_letter(col_idx)
                            range_ref = f"{col_letter}2:{col_letter}{len(df) + 1}"

                            color_scale = ColorScaleRule(
                                start_type='min', start_color=rule_config.get('start_color', 'FF0000'),
                                end_type='max', end_color=rule_config.get('end_color', '00FF00')
                            )
                            worksheet.conditional_formatting.add(range_ref, color_scale)
            except Exception as e:
                logger.warning(f"Failed to apply conditional formatting rule {rule_name}: {e}")

    @staticmethod
    def _apply_workbook_formatting(workbook: Workbook) -> None:
        """Apply workbook-level formatting."""
        # Set default font for all sheets
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.font == Font():  # Default font
                        cell.font = Font(name='Calibri', size=11)

    @staticmethod
    def _get_file_size_mb(file_path: str) -> float:
        """Get file size in megabytes."""
        try:
            return Path(file_path).stat().st_size / (1024 * 1024)
        except TypeError:
            return 0.0