"""
Output Service - Handles Excel file generation and formatting
"""

import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dataclasses import asdict

logger = logging.getLogger(__name__)

class ExcelOutputService:
    """Handles Excel file generation with proper formatting"""

    def __init__(self, output_file_path: str):
        self.output_file_path = output_file_path
        self.sheets_config = self._get_sheets_configuration()

    def _get_sheets_configuration(self) -> Dict[str, Dict]:
        """Define configuration for each Excel sheet"""
        return {
            "Correct Applications": {
                "description": "Valid, correctly formatted applications",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "YearStart", "YearEnd", "Make", "Code",
                           "Model", "Note", "Original"]
            },
            "Incorrect Applications": {
                "description": "Applications with format issues",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "YearStart", "YearEnd", "Make", "Code",
                           "Model", "Note", "Original"]
            },
            "Invalid Applications": {
                "description": "Applications that failed validation",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Line", "Reason"]
            },
            "Invalid Years": {
                "description": "Applications with invalid year ranges",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Line", "Reason"]
            },
            "Illegal Characters": {
                "description": "Applications containing illegal characters",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "OriginalPartApplication", "IllegalCharacters"]
            },
            "Date Corrections": {
                "description": "Date format corrections applied",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "OriginalDate", "CorrectedDate", "Note"]
            },
            "Discrepancies": {
                "description": "Data discrepancies found during validation",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Make", "Code", "Model", "Field", "Original", "Corrected"]
            },
            "Validated Notes": {
                "description": "Note validation results",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Note", "Notes_Validity"]
            },
            "Key Occurrences": {
                "description": "Lookup key usage statistics",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["Key", "Occurrences"]
            },
            "Unique Words": {
                "description": "Unique words found in notes",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["Word", "PartNumbers"]
            },
            "Application Review": {
                "description": "Applications formatted for review",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Make", "Code", "Model", "Year", "Original Note",
                           "Note", "Liter", "LHD", "RHD", "Front Brake", "Rear Brake",
                           "Manual Transmission", "Automatic Transmission", "Transmission",
                           "Front Axle", "Rear Axle", "Fuel", "Doors"]
            },
            "Expanded Applications": {
                "description": "Applications expanded by year",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Make", "Code", "Model", "Year", "Original Note",
                           "Note", "Liter", "LHD", "RHD", "Front Brake", "Rear Brake",
                           "Manual Transmission", "Automatic Transmission", "Transmission",
                           "Front Axle", "Rear Axle", "Fuel", "Doors"]
            },
            "Jeep Applications": {
                "description": "Applications specific to Jeep vehicles",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Application"]
            },
            "All Applications": {
                "description": "All reconstructed applications",
                "freeze_row": 1,
                "auto_filter": True,
                "columns": ["PartNumber", "Application"]
            }
        }

    def generate_excel_report(self, results_data: Dict[str, List[Any]]) -> bool:
        """Generate comprehensive Excel report"""
        try:
            # Convert data to DataFrames
            dataframes = self._prepare_dataframes(results_data)

            # Write to Excel with multiple sheets
            with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    if not df.empty:
                        df.to_excel(writer, index=False, sheet_name=sheet_name)
                        logger.info(f"Written {len(df)} rows to sheet '{sheet_name}'")

            # Apply formatting
            self._apply_excel_formatting()

            logger.info(f"Excel report generated: {self.output_file_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return False

    def _prepare_dataframes(self, results_data: Dict[str, List[Any]]) -> Dict[str, pd.DataFrame]:
        """Convert results data to pandas DataFrames"""
        dataframes = {}

        # Standard data conversions
        simple_conversions = [
            "invalid_lines", "invalid_years", "illegal_characters",
            "date_corrections", "validated_notes"
        ]

        for key in simple_conversions:
            if key in results_data and results_data[key]:
                sheet_name = self._get_sheet_name(key)
                dataframes[sheet_name] = pd.DataFrame(results_data[key])

        # Convert application objects to DataFrames
        if "correct_applications" in results_data and results_data["correct_applications"]:
            correct_df = self._applications_to_dataframe(results_data["correct_applications"])
            if not correct_df.empty:
                dataframes["Correct Applications"] = correct_df

        if "incorrect_applications" in results_data and results_data["incorrect_applications"]:
            incorrect_df = self._applications_to_dataframe(results_data["incorrect_applications"])
            if not incorrect_df.empty:
                dataframes["Incorrect Applications"] = incorrect_df

        # Key occurrences (from lookup service)
        if "key_occurrences" in results_data and results_data["key_occurrences"]:
            key_data = [
                {"Key": key, "Occurrences": count}
                for key, count in results_data["key_occurrences"].items()
            ]
            if key_data:
                df = pd.DataFrame(key_data).sort_values(by=["Occurrences", "Key"],
                                                       ascending=[False, True])
                dataframes["Key Occurrences"] = df

        # Unique words analysis
        if "unique_words" in results_data and results_data["unique_words"]:
            df = pd.DataFrame(results_data["unique_words"])
            if not df.empty:
                df = df.sort_values(by="Word").reset_index(drop=True)
                dataframes["Unique Words"] = df

        # Reconstructed applications
        if "all_applications" in results_data and results_data["all_applications"]:
            dataframes["All Applications"] = pd.DataFrame(results_data["all_applications"])

        if "jeep_applications" in results_data and results_data["jeep_applications"]:
            dataframes["Jeep Applications"] = pd.DataFrame(results_data["jeep_applications"])

        # Expanded applications (by year)
        if "expanded_applications" in results_data and results_data["expanded_applications"]:
            expanded_df = pd.DataFrame(results_data["expanded_applications"])
            if not expanded_df.empty:
                # Sort by part number and year
                expanded_df = expanded_df.sort_values(by=["PartNumber", "Year"])
                dataframes["Expanded Applications"] = expanded_df

        # Application review format
        if "application_review" in results_data and results_data["application_review"]:
            review_df = pd.DataFrame(results_data["application_review"])
            if not review_df.empty:
                dataframes["Application Review"] = review_df

        # Create discrepancies sheet if we have verification data
        # This would be populated by a verification service (not implemented in this refactor)
        if "discrepancies" in results_data and results_data["discrepancies"]:
            dataframes["Discrepancies"] = pd.DataFrame(results_data["discrepancies"])

        return dataframes

    def _applications_to_dataframe(self, applications: List[Any]) -> pd.DataFrame:
        """Convert application objects to DataFrame"""
        if not applications:
            return pd.DataFrame()

        # Convert dataclass objects to dictionaries
        if hasattr(applications[0], '__dataclass_fields__'):
            data = [asdict(app) for app in applications]
        else:
            data = applications

        df = pd.DataFrame(data)

        # Clean string data
        string_columns = df.select_dtypes(include=['object']).columns
        for col in string_columns:
            df[col] = df[col].apply(self._clean_string)

        return df

    def _add_specialized_sheets(self, dataframes: Dict[str, pd.DataFrame],
                              results_data: Dict[str, List[Any]]):
        """Add specialized analysis sheets"""

        # Key occurrences (from lookup service)
        if "key_occurrences" in results_data:
            key_data = [
                {"Key": key, "Occurrences": count}
                for key, count in results_data["key_occurrences"].items()
            ]
            if key_data:
                df = pd.DataFrame(key_data).sort_values(by=["Key", "Occurrences"],
                                                       ascending=[True, False])
                dataframes["Key Occurrences"] = df

        # Unique words analysis
        if "unique_words" in results_data:
            dataframes["Unique Words"] = pd.DataFrame(results_data["unique_words"])

        # Reconstructed applications
        if "all_applications" in results_data:
            dataframes["All Applications"] = pd.DataFrame(results_data["all_applications"])

        if "jeep_applications" in results_data:
            dataframes["Jeep Applications"] = pd.DataFrame(results_data["jeep_applications"])

        # Expanded applications (by year)
        if "expanded_applications" in results_data:
            expanded_df = self._create_expanded_applications(results_data["correct_applications"])
            if not expanded_df.empty:
                dataframes["Expanded Applications"] = expanded_df

        # Application review format
        if "application_review" in results_data:
            review_df = self._create_review_format(results_data["correct_applications"])
            if not review_df.empty:
                dataframes["Application Review"] = review_df

    def _create_expanded_applications(self, applications: List[Any]) -> pd.DataFrame:
        """Create year-expanded applications"""
        expanded_data = []

        for app in applications:
            if hasattr(app, 'year_start') and hasattr(app, 'year_end'):
                try:
                    start_year = int(app.year_start) if app.year_start else 0
                    end_year = int(app.year_end) if app.year_end else 0

                    if start_year > 0 and end_year > 0:
                        for year in range(start_year, end_year + 1):
                            expanded_row = asdict(app)
                            expanded_row["Year"] = year
                            # Add attribute columns with defaults
                            attribute_columns = [
                                "Liter", "LHD", "RHD", "Front Brake", "Rear Brake",
                                "Manual Transmission", "Automatic Transmission",
                                "Transmission", "Front Axle", "Rear Axle", "Fuel", "Doors"
                            ]
                            for col in attribute_columns:
                                if col not in expanded_row:
                                    expanded_row[col] = ""

                            expanded_data.append(expanded_row)
                except (ValueError, TypeError):
                    continue

        return pd.DataFrame(expanded_data) if expanded_data else pd.DataFrame()

    def _create_review_format(self, applications: List[Any]) -> pd.DataFrame:
        """Create application review format"""
        review_data = []

        for app in applications:
            review_row = asdict(app) if hasattr(app, '__dataclass_fields__') else app

            # Add required columns for review format
            review_columns = [
                "PartNumber", "Make", "Code", "Model", "Year", "Original Note",
                "Note", "Liter", "LHD", "RHD", "Front Brake", "Rear Brake",
                "Manual Transmission", "Automatic Transmission", "Transmission",
                "Front Axle", "Rear Axle", "Fuel", "Doors"
            ]

            formatted_row = {}
            for col in review_columns:
                if col in review_row:
                    formatted_row[col] = review_row[col]
                elif col == "Original Note":
                    formatted_row[col] = review_row.get("note", "")
                else:
                    formatted_row[col] = ""

            review_data.append(formatted_row)

        return pd.DataFrame(review_data) if review_data else pd.DataFrame()

    def _apply_excel_formatting(self):
        """Apply formatting to Excel file"""
        try:
            workbook = load_workbook(self.output_file_path)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                config = self.sheets_config.get(sheet_name, {})

                # Freeze header row
                if config.get("freeze_row", 0) > 0:
                    sheet.freeze_panes = f"A{config['freeze_row'] + 1}"

                # Apply auto filter
                if config.get("auto_filter", False) and sheet.max_row > 1:
                    max_column = sheet.max_column
                    max_col_letter = get_column_letter(max_column)
                    sheet.auto_filter.ref = f"A1:{max_col_letter}1"

                # Auto-adjust column widths (basic implementation)
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    # Set column width with reasonable limits
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(self.output_file_path)
            logger.info("Excel formatting applied successfully")

        except Exception as e:
            logger.error(f"Failed to apply Excel formatting: {e}")

    def _get_sheet_name(self, data_key: str) -> str:
        """Map data key to sheet name"""
        mapping = {
            "invalid_lines": "Invalid Applications",
            "invalid_years": "Invalid Years",
            "illegal_characters": "Illegal Characters",
            "date_corrections": "Date Corrections",
            "discrepancies": "Discrepancies",
            "validated_notes": "Validated Notes"
        }
        return mapping.get(data_key, data_key.replace("_", " ").title())

    def _clean_string(self, value) -> str:
        """Clean string values of non-printable characters"""
        if not isinstance(value, str):
            return value

        import string
        return ''.join(char for char in value if char in string.printable)

    def generate_summary_report(self, results_data: Dict[str, List[Any]]) -> str:
        """Generate a text summary of processing results"""
        summary_lines = []
        summary_lines.append("=" * 60)
        summary_lines.append("APPLICATION PROCESSING SUMMARY")
        summary_lines.append("=" * 60)

        # Count statistics
        correct_count = len(results_data.get("correct_applications", []))
        incorrect_count = len(results_data.get("incorrect_applications", []))
        invalid_count = len(results_data.get("invalid_lines", []))

        summary_lines.append(f"Correct Applications: {correct_count:,}")
        summary_lines.append(f"Incorrect Applications: {incorrect_count:,}")
        summary_lines.append(f"Invalid Applications: {invalid_count:,}")
        summary_lines.append(f"Total Processed: {correct_count + incorrect_count + invalid_count:,}")

        # Error statistics
        year_errors = len(results_data.get("invalid_years", []))
        char_errors = len(results_data.get("illegal_characters", []))
        discrepancies = len(results_data.get("discrepancies", []))

        summary_lines.append("")
        summary_lines.append("ERROR SUMMARY:")
        summary_lines.append(f"  Invalid Years: {year_errors:,}")
        summary_lines.append(f"  Illegal Characters: {char_errors:,}")
        summary_lines.append(f"  Data Discrepancies: {discrepancies:,}")

        # Key usage statistics
        key_occurrences = results_data.get("key_occurrences", {})
        if key_occurrences:
            used_keys = sum(1 for count in key_occurrences.values() if count > 0)
            total_keys = len(key_occurrences)

            summary_lines.append("")
            summary_lines.append("LOOKUP KEY USAGE:")
            summary_lines.append(f"  Keys Used: {used_keys}/{total_keys}")
            summary_lines.append(f"  Usage Rate: {used_keys/total_keys*100:.1f}%")

        summary_lines.append("=" * 60)
        summary_lines.append(f"Report generated: {self.output_file_path}")
        summary_lines.append("=" * 60)

        return "\n".join(summary_lines)

# ==============================================================================
# Output Service Factory
# ==============================================================================

class OutputServiceFactory:
    """Factory for creating different types of output services"""

    @staticmethod
    def create_excel_service(output_path: str) -> ExcelOutputService:
        """Create Excel output service"""
        return ExcelOutputService(output_path)

    @staticmethod
    def create_csv_service(output_dir: str):
        """Create CSV output service (future implementation)"""
        # Could implement CSV output for different use cases
        pass

    @staticmethod
    def create_json_service(output_path: str):
        """Create JSON output service (future implementation)"""
        # Could implement JSON output for API consumption
        pass