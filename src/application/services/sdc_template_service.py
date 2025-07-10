# src/application/services/sdc_template_service.py
"""
SDC template service with marketing description integration.
"""

import logging
from typing import List, Dict, Any, Optional
from pathlib import Path

from ...domain.models import ProcessingResult
from ...infrastructure.repositories.filemaker.marketing_description_repository import FilemakerMarketingDescriptionRepository
from .marketing_description_service import MarketingDescriptionService

logger = logging.getLogger(__name__)


class SdcTemplateService:
    """Service for SDC template generation with marketing descriptions."""

    def __init__(
            self,
            filemaker_repository: FilemakerMarketingDescriptionRepository,
            marketing_service: MarketingDescriptionService
    ):
        self.filemaker_repository = filemaker_repository
        self.marketing_service = marketing_service

    def generate_sdc_template(
            self,
            template_file: str,
            output_file: str,
            missing_parts_file: Optional[str] = None
    ) -> ProcessingResult:
        """Generate SDC template with marketing descriptions."""
        logger.info(f"Generating SDC template: {template_file} -> {output_file}")

        try:
            # Get master data with marketing descriptions
            master_data = self.filemaker_repository.get_master_data_with_descriptions()

            # Filter by missing parts if specified
            if missing_parts_file and Path(missing_parts_file).exists():
                missing_parts = self._load_missing_parts(missing_parts_file)
                master_data = [
                    record for record in master_data
                    if record.get('AS400_NumberStripped') in missing_parts
                ]

            # Load Excel template and populate
            populated_count = self._populate_template(template_file, output_file, master_data)

            return ProcessingResult(
                success=True,
                items_processed=populated_count,
                data={
                    'input_file': template_file,
                    'output_file': output_file,
                    'total_products': len(master_data),
                    'populated_products': populated_count
                }
            )

        except Exception as e:
            logger.error(f"SDC template generation failed: {e}")
            return ProcessingResult(
                success=False,
                errors=[f"SDC template generation failed: {e}"]
            )

    def _load_missing_parts(self, missing_parts_file: str) -> set:
        """Load missing part numbers from file."""
        missing_parts = set()
        try:
            with open(missing_parts_file, 'r') as f:
                for line in f:
                    part_number = line.strip()
                    if part_number:
                        missing_parts.add(part_number)
        except Exception as e:
            logger.warning(f"Failed to load missing parts file: {e}")

        return missing_parts

    def _populate_template(self, template_file: str, output_file: str, master_data: List[Dict[str, Any]]) -> int:
        """Populate Excel template with master data."""
        from openpyxl import load_workbook

        wb = load_workbook(filename=template_file)

        # Populate PiesDescriptions sheet with marketing descriptions
        populated_count = self._populate_pies_descriptions(wb, master_data)

        # Save the populated template
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)

        logger.info(f"SDC template populated with {populated_count} products")
        return populated_count

    def _populate_pies_descriptions(self, workbook, master_data: List[Dict[str, Any]]) -> int:
        """Populate PiesDescriptions sheet with marketing descriptions."""
        ws = workbook['PiesDescriptions']
        row_num = 2  # Start after header
        populated_count = 0

        for record in master_data:
            if record.get('PartBrand') == 'RT Off-Road':
                continue  # Skip RT Off-Road products

            row_num += 1
            populated_count += 1

            try:
                # Basic product information
                ws[f"A{row_num}"] = 'Default'
                ws[f"B{row_num}"] = record.get('SDC_PartType', '')
                ws[f"C{row_num}"] = record.get('AS400_NumberStripped', '')
                ws[f"D{row_num}"] = 'EN'

                # Marketing description with fallback logic
                terminology_id = record.get('SDC_PartTerminologyID', '')
                fallback_description = record.get('RTOffRoadAdCopy', '')

                marketing_description = self.marketing_service.get_description_for_sdc(
                    terminology_id, fallback_description
                )

                # Column P is for marketing description
                ws[f"P{row_num}"] = marketing_description

                # Other description fields
                ws[f"E{row_num}"] = record.get('SDC_DescriptionAbbreviated', '')[:12]
                ws[f"F{row_num}"] = record.get('SDC_DescriptionShort', '')[:20]
                ws[f"G{row_num}"] = record.get('SDC_DescriptionInvoice', '')[:40]

            except Exception as e:
                logger.error(f"Error populating row {row_num}: {e}")
                continue

        return populated_count